import os
import uuid
import logging
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import List, Optional

from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.security import OAuth2PasswordBearer
from starlette.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from jose import jwt, JWTError
from pydantic import BaseModel, EmailStr, Field
import openpyxl
import io
import base64
from PIL import Image

from whatsapp_service import send_whatsapp, build_assignment_message
from pdf_service import build_orden_pdf
from suministros_module import build_router as build_suministros_router, init_suministros
from email_service import send_email, build_welcome_email
from geocoding_service import geocode_address, haversine, nearest_neighbor_sort
from fastapi.responses import StreamingResponse

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# ----------------- Config -----------------
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = os.environ.get("JWT_ALGORITHM", "HS256")
JWT_EXPIRES_MINUTES = int(os.environ.get("JWT_EXPIRES_MINUTES", "1440"))
ADMIN_EMAIL = os.environ["ADMIN_EMAIL"]
ADMIN_PASSWORD = os.environ["ADMIN_PASSWORD"]
ADMIN_NOMBRE = os.environ.get("ADMIN_NOMBRE", "Admin")
ADMIN_APELLIDOS = os.environ.get("ADMIN_APELLIDOS", "MVG")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# ----------------- DB -----------------
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

users_col = db.users
clientes_col = db.clientes
sucursales_col = db.sucursales
ordenes_col = db.ordenes


# ----------------- Image compression -----------------
def compress_evidencia(b64: Optional[str], max_side: int = 1100, quality: int = 55) -> Optional[str]:
    """Compress a base64 image (with or without data URL prefix) to keep
    each pin pad photo well under 1MB. Returns a data URL on success
    or the original string if compression fails (best-effort)."""
    if not b64 or not isinstance(b64, str):
        return b64
    try:
        data_str = b64
        if data_str.startswith("data:"):
            _, _, rest = data_str.partition(",")
            data_str = rest
        raw = base64.b64decode(data_str)
        # Cheap guard: if already small (<300KB), keep original to avoid recompression artifacts
        if len(raw) < 300 * 1024:
            return b64 if b64.startswith("data:") else f"data:image/jpeg;base64,{data_str}"
        img = Image.open(io.BytesIO(raw))
        # Convert (RGBA / P) to RGB JPEG
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        # Resize if too big
        w, h = img.size
        if max(w, h) > max_side:
            ratio = max_side / float(max(w, h))
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=quality, optimize=True, progressive=True)
        return "data:image/jpeg;base64," + base64.b64encode(out.getvalue()).decode("ascii")
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.warning("[compress_evidencia] failed: %s", e)
        return b64


# ----------------- Auth utils -----------------
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")


def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return pwd_context.verify(plain, hashed)
    except Exception:
        return False


def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=JWT_EXPIRES_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ----------------- Models -----------------
class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict


class UserPublic(BaseModel):
    id: str
    email: EmailStr
    role: str
    nombre: str
    apellidos: str
    rut: Optional[str] = None
    telefono: Optional[str] = None


class TecnicoCreate(BaseModel):
    rut: str
    nombre: str
    apellidos: str
    email: EmailStr
    telefono: str
    password: str = Field(min_length=6)
    bodega_id: Optional[str] = None
    direccion: Optional[str] = None
    comuna: Optional[str] = None
    region: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None


class TecnicoUpdate(BaseModel):
    rut: Optional[str] = None
    nombre: Optional[str] = None
    apellidos: Optional[str] = None
    email: Optional[EmailStr] = None
    telefono: Optional[str] = None
    password: Optional[str] = None
    bodega_id: Optional[str] = None
    direccion: Optional[str] = None
    comuna: Optional[str] = None
    region: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None


class ClienteCreate(BaseModel):
    nombre: str
    nombre_fantasia: Optional[str] = None
    rut: Optional[str] = None
    contacto: Optional[str] = None
    email: Optional[EmailStr] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None


class ClienteUpdate(BaseModel):
    nombre: Optional[str] = None
    nombre_fantasia: Optional[str] = None
    rut: Optional[str] = None
    contacto: Optional[str] = None
    email: Optional[EmailStr] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None


class Cliente(BaseModel):
    id: str
    nombre: str
    nombre_fantasia: Optional[str] = None
    rut: Optional[str] = None
    contacto: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    direccion: Optional[str] = None
    created_at: str


class SucursalCreate(BaseModel):
    cliente_id: str
    nombre: str
    codigo_comercio: Optional[str] = None
    direccion: Optional[str] = None
    comuna: Optional[str] = None
    region: Optional[str] = None
    telefono: Optional[str] = None
    encargado: Optional[str] = None


class SucursalUpdate(BaseModel):
    nombre: Optional[str] = None
    codigo_comercio: Optional[str] = None
    direccion: Optional[str] = None
    comuna: Optional[str] = None
    region: Optional[str] = None
    telefono: Optional[str] = None
    encargado: Optional[str] = None


class OrdenCreate(BaseModel):
    cliente_id: str
    sucursal_id: str
    tecnico_id: Optional[str] = None
    titulo: str
    descripcion: str
    prioridad: str = Field(pattern="^(baja|media|alta)$")
    serie: Optional[str] = None
    modelo: Optional[str] = None
    ddll: Optional[str] = None
    fecha_limite: Optional[str] = None  # ISO date "YYYY-MM-DD"


class OrdenUpdate(BaseModel):
    titulo: Optional[str] = None
    descripcion: Optional[str] = None
    prioridad: Optional[str] = Field(default=None, pattern="^(baja|media|alta)$")
    tecnico_id: Optional[str] = None
    serie: Optional[str] = None
    modelo: Optional[str] = None
    ddll: Optional[str] = None
    fecha_limite: Optional[str] = None
    fecha_ejecucion: Optional[str] = None  # YYYY-MM-DD o ISO datetime


class OrdenAsignar(BaseModel):
    tecnico_id: str


class OrdenFinalizar(BaseModel):
    evidencia_base64: str  # base64 image data
    notas: Optional[str] = None
    # Location is required - técnico must authorize browser geolocation
    lat: float
    lng: float
    address: Optional[str] = None
    accuracy_m: Optional[float] = None


class MaterialUsadoItem(BaseModel):
    sku: str
    descripcion: Optional[str] = None
    cantidad: int


class PinPadUpdate(BaseModel):
    # Foto principal (compat con flujo anterior). Si se reciben los nuevos
    # campos foto_antes/foto_despues, este queda como alias del "antes" o
    # "después" (se mantiene por retrocompatibilidad).
    evidencia_base64: Optional[str] = None

    # NUEVOS campos (4 fotos protocolo MVG)
    foto_antes_base64: Optional[str] = None             # 1) Pinpad ANTES de actualizar (OBLIGATORIA)
    foto_descarga_master_base64: Optional[str] = None   # 2) Informe Descarga Master (OPCIONAL)
    foto_despues_base64: Optional[str] = None           # 3) Pinpad DESPUÉS de actualizar (OBLIGATORIA)
    foto_comprobante_venta_base64: Optional[str] = None # 4) Comprobante de venta (OPCIONAL)

    notas: Optional[str] = None
    lat: Optional[float] = None
    lng: Optional[float] = None
    address: Optional[str] = None
    accuracy_m: Optional[float] = None
    materiales_usados: Optional[List[MaterialUsadoItem]] = None
    sin_suministros: Optional[bool] = False


class PinPadEditPhoto(BaseModel):
    """Edición de fotos individuales del pinpad. Permite actualizar
    cualquiera de las 4 fotos. Mantiene compat con el flujo viejo."""
    evidencia_base64: Optional[str] = None
    foto_antes_base64: Optional[str] = None
    foto_descarga_master_base64: Optional[str] = None
    foto_despues_base64: Optional[str] = None
    foto_comprobante_venta_base64: Optional[str] = None
    notas: Optional[str] = None


# ----------------- Disponibilidad técnico -----------------
class DisponibilidadDia(BaseModel):
    activo: bool = False
    hora_inicio: str = "09:00"   # formato HH:MM (24h), entre 00:00 y 23:59
    hora_fin: str = "20:00"


class DisponibilidadSemanal(BaseModel):
    """Disponibilidad semanal del técnico. Claves: lun, mar, mié, jue, vie, sáb, dom."""
    lun: DisponibilidadDia = DisponibilidadDia()
    mar: DisponibilidadDia = DisponibilidadDia()
    mie: DisponibilidadDia = DisponibilidadDia()
    jue: DisponibilidadDia = DisponibilidadDia()
    vie: DisponibilidadDia = DisponibilidadDia()
    sab: DisponibilidadDia = DisponibilidadDia()
    dom: DisponibilidadDia = DisponibilidadDia()


class OrdenCUEUpload(BaseModel):
    """Subida de foto del CUE (Comprobante Único Electrónico) para una orden."""
    cue_base64: str
    notas: Optional[str] = None


# ----------------- Auth dependencies -----------------
async def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="No autorizado",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id: str = payload.get("sub")
        if not user_id:
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    user = await users_col.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
    if not user:
        raise credentials_exception
    return user


async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
    return user


async def require_tecnico(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "tecnico":
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
    return user


# ----------------- App & Router -----------------
app = FastAPI(title="MVG Computación API")
api_router = APIRouter(prefix="/api")


# ----------------- Helpers -----------------
def clean_user(u: dict) -> dict:
    if not u:
        return u
    u = {k: v for k, v in u.items() if k not in ("_id", "hashed_password")}
    return u


async def enrich_orden(o: dict) -> dict:
    """Add cliente, sucursal, tecnico info to orden."""
    cliente = await clientes_col.find_one({"id": o.get("cliente_id")}, {"_id": 0})
    sucursal = await sucursales_col.find_one({"id": o.get("sucursal_id")}, {"_id": 0})
    tecnico = await users_col.find_one(
        {"id": o.get("tecnico_id")}, {"_id": 0, "hashed_password": 0}
    )
    return {
        **{k: v for k, v in o.items() if k != "_id"},
        "cliente": cliente,
        "sucursal": sucursal,
        "tecnico": tecnico,
    }


# ----------------- Auth Routes -----------------
@api_router.post("/auth/login", response_model=TokenResponse)
async def login(payload: LoginRequest):
    user = await users_col.find_one({"email": payload.email.lower()})
    if not user or not verify_password(payload.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    token = create_access_token({"sub": user["id"], "role": user["role"]})
    return TokenResponse(access_token=token, user=clean_user(user))


@api_router.get("/auth/me")
async def me(current: dict = Depends(get_current_user)):
    return current


# ----------------- Admin: Tecnicos -----------------
@api_router.post("/admin/tecnicos", status_code=201)
async def create_tecnico(payload: TecnicoCreate, _: dict = Depends(require_admin)):
    existing = await users_col.find_one({"email": payload.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email ya registrado")
    rut_existing = await users_col.find_one({"rut": payload.rut})
    if rut_existing:
        raise HTTPException(status_code=400, detail="RUT ya registrado")

    # Auto-geocode si tenemos dirección pero faltan coords
    lat = payload.lat
    lng = payload.lng
    if (lat is None or lng is None) and payload.direccion:
        try:
            result = await geocode_address(
                db, payload.direccion, payload.comuna, payload.region
            )
            if result:
                lat, lng, _ = result
                logger.info(
                    "[geocode tecnico create] %s -> (%s, %s)",
                    payload.direccion, lat, lng,
                )
        except Exception as e:
            logger.warning("[geocode tecnico create] err: %s", e)

    new_user = {
        "id": str(uuid.uuid4()),
        "email": payload.email.lower(),
        "rut": payload.rut,
        "nombre": payload.nombre,
        "apellidos": payload.apellidos,
        "telefono": payload.telefono,
        "role": "tecnico",
        "hashed_password": hash_password(payload.password),
        "bodega_id": payload.bodega_id,
        "direccion": payload.direccion,
        "comuna": payload.comuna,
        "region": payload.region,
        "lat": lat,
        "lng": lng,
        "created_at": now_iso(),
    }
    await users_col.insert_one(new_user)

    # Send welcome email (best-effort, non-blocking on failure)
    welcome_result = None
    try:
        bodega_label = ""
        region_label = ""
        if payload.bodega_id:
            bodega_doc = await db.bodegas.find_one({"id": payload.bodega_id})
            if bodega_doc:
                bodega_label = bodega_doc.get("nombre") or ""
                region_label = bodega_doc.get("region") or ""
        nombre_completo = f"{payload.nombre} {payload.apellidos}".strip()
        html, text = build_welcome_email(
            tecnico_nombre=nombre_completo,
            tecnico_email=payload.email,
            plain_password=payload.password,
            bodega=bodega_label,
            region=region_label,
        )
        welcome_result = await send_email(
            to=[payload.email],
            subject=f"Bienvenido a MVG Computación - {nombre_completo}",
            html=html,
            text=text,
        )
        logger.info("[Welcome email] %s -> %s", payload.email, welcome_result.get("mode"))
    except Exception as e:
        logger.warning("[Welcome email] error: %s", e)
        welcome_result = {"mode": "error", "error": str(e)}

    user = clean_user(new_user)
    user["welcome_email"] = welcome_result

    # Send credentials via WhatsApp (best-effort)
    wa_result = None
    if payload.telefono:
        try:
            nombre_completo = f"{payload.nombre} {payload.apellidos}".strip()
            msg = (
                f"🔐 *MVG Computación - Tu cuenta*\n\n"
                f"Hola {nombre_completo},\n\n"
                f"Tu cuenta ha sido creada. Estas son tus credenciales:\n\n"
                f"📧 Email: {payload.email}\n"
                f"🔑 Contraseña: *{payload.password}*\n\n"
                f"🌐 Accede en: https://mvg.goroky.es\n\n"
                f"⚠️ Por seguridad, cambia tu contraseña al iniciar sesión.\n"
                f"Revisa la plataforma varias veces al día para mantener "
                f"tus órdenes al día."
            )
            wa_result = await send_whatsapp(payload.telefono, msg)
            logger.info(
                "[WhatsApp create_tecnico] %s -> %s",
                payload.telefono,
                wa_result.get("ok") if isinstance(wa_result, dict) else "err",
            )
        except Exception as e:
            logger.warning("[WhatsApp create_tecnico] error: %s", e)
            wa_result = {"ok": False, "error": str(e)}
    user["whatsapp_welcome"] = wa_result
    return user


class PasswordWhatsApp(BaseModel):
    password: str = Field(min_length=6)


@api_router.post("/admin/tecnicos/{tecnico_id}/enviar-password-whatsapp")
async def enviar_password_whatsapp(
    tecnico_id: str,
    payload: PasswordWhatsApp,
    _: dict = Depends(require_admin),
):
    """Updates the technician's password AND sends it via WhatsApp.

    The plain-text password is sent via WhatsApp ONCE in this message; the
    DB only stores the bcrypt hash. The admin is responsible for ensuring
    the channel is appropriate.
    """
    tec = await users_col.find_one({"id": tecnico_id, "role": "tecnico"})
    if not tec:
        raise HTTPException(status_code=404, detail="Técnico no encontrado")
    if not tec.get("telefono"):
        raise HTTPException(
            status_code=400, detail="El técnico no tiene teléfono registrado"
        )
    # Update password
    await users_col.update_one(
        {"id": tecnico_id},
        {"$set": {"hashed_password": hash_password(payload.password)}},
    )

    nombre = f"{tec.get('nombre','')} {tec.get('apellidos','')}".strip()
    msg = (
        f"🔐 *MVG Computación - Credenciales*\n\n"
        f"Hola {nombre},\n\n"
        f"Tu nueva contraseña de acceso es:\n"
        f"*{payload.password}*\n\n"
        f"📧 Email: {tec.get('email')}\n\n"
        f"⚠️ Por seguridad, cámbiala tras iniciar sesión.\n"
        f"Mantén tus datos seguros y revisa la plataforma "
        f"varias veces al día para mantener tus órdenes al día."
    )
    wa_result = await send_whatsapp(tec["telefono"], msg)
    return {"ok": True, "whatsapp": wa_result}


@api_router.post("/admin/tecnicos/{tecnico_id}/impersonate")
async def impersonate_tecnico(
    tecnico_id: str, admin: dict = Depends(require_admin)
):
    """Returns a JWT token to login AS the técnico. Admin retains their own
    session in the client; this is for previewing the técnico's app."""
    tec = await users_col.find_one({"id": tecnico_id, "role": "tecnico"})
    if not tec:
        raise HTTPException(status_code=404, detail="Técnico no encontrado")
    token = create_access_token(
        {"sub": tec["id"], "role": tec["role"], "impersonated_by": admin["id"]}
    )
    return {"access_token": token, "token_type": "bearer", "tecnico": clean_user(tec)}


@api_router.get("/admin/tecnicos")
async def list_tecnicos(_: dict = Depends(require_admin)):
    cursor = users_col.find(
        {"role": "tecnico"}, {"_id": 0, "hashed_password": 0}
    ).sort("created_at", -1)
    return await cursor.to_list(1000)


@api_router.patch("/admin/tecnicos/{tecnico_id}")
async def update_tecnico(
    tecnico_id: str, payload: TecnicoUpdate, _: dict = Depends(require_admin)
):
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "password" in update_data:
        update_data["hashed_password"] = hash_password(update_data.pop("password"))
    if "email" in update_data:
        update_data["email"] = update_data["email"].lower()
    if not update_data:
        raise HTTPException(status_code=400, detail="Sin cambios")

    # Si dirección/comuna/region cambia y NO se pasan lat/lng explícitos,
    # re-geocodificar
    dir_changed = any(k in update_data for k in ("direccion", "comuna", "region"))
    coords_explicit = "lat" in update_data and "lng" in update_data
    if dir_changed and not coords_explicit:
        existing = await users_col.find_one({"id": tecnico_id, "role": "tecnico"})
        if existing:
            new_dir = update_data.get("direccion", existing.get("direccion"))
            new_com = update_data.get("comuna", existing.get("comuna"))
            new_reg = update_data.get("region", existing.get("region"))
            if new_dir:
                try:
                    r = await geocode_address(db, new_dir, new_com, new_reg)
                    if r:
                        update_data["lat"], update_data["lng"], _ = r
                        logger.info(
                            "[geocode tec update] %s -> (%s, %s)",
                            new_dir, update_data["lat"], update_data["lng"],
                        )
                    else:
                        update_data["lat"] = None
                        update_data["lng"] = None
                except Exception as e:
                    logger.warning("[geocode tec update] err: %s", e)

    result = await users_col.update_one(
        {"id": tecnico_id, "role": "tecnico"}, {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Técnico no encontrado")
    user = await users_col.find_one(
        {"id": tecnico_id}, {"_id": 0, "hashed_password": 0}
    )
    return user


# ----------------- Geocoding (admin helper para UI) -----------------

class GeocodePayload(BaseModel):
    direccion: str
    comuna: Optional[str] = None
    region: Optional[str] = None


@api_router.post("/admin/geocode")
async def admin_geocode(payload: GeocodePayload, _: dict = Depends(require_admin)):
    """Convierte dirección textual en coordenadas (usa Nominatim/OSM).
    Útil para previsualizar en el formulario de técnico antes de guardar."""
    if not payload.direccion or len(payload.direccion.strip()) < 4:
        raise HTTPException(status_code=400, detail="Dirección demasiado corta")
    result = await geocode_address(
        db, payload.direccion, payload.comuna, payload.region
    )
    if not result:
        return {"ok": False, "detail": "No se pudo geocodificar"}
    lat, lng, display = result
    return {"ok": True, "lat": lat, "lng": lng, "display_name": display}


@api_router.delete("/admin/tecnicos/{tecnico_id}")
async def delete_tecnico(tecnico_id: str, _: dict = Depends(require_admin)):
    result = await users_col.delete_one({"id": tecnico_id, "role": "tecnico"})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Técnico no encontrado")
    return {"ok": True}


# ----------------- Admin: Clientes -----------------
@api_router.post("/admin/clientes", status_code=201)
async def create_cliente(payload: ClienteCreate, _: dict = Depends(require_admin)):
    doc = {
        "id": str(uuid.uuid4()),
        **payload.model_dump(),
        "created_at": now_iso(),
    }
    await clientes_col.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api_router.get("/admin/clientes")
async def list_clientes(_: dict = Depends(require_admin)):
    cursor = clientes_col.find({}, {"_id": 0}).sort("created_at", -1)
    return await cursor.to_list(1000)


@api_router.get("/admin/clientes/{cliente_id}")
async def get_cliente(cliente_id: str, _: dict = Depends(require_admin)):
    cliente = await clientes_col.find_one({"id": cliente_id}, {"_id": 0})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    sucursales = await sucursales_col.find(
        {"cliente_id": cliente_id}, {"_id": 0}
    ).to_list(1000)
    return {**cliente, "sucursales": sucursales}


@api_router.patch("/admin/clientes/{cliente_id}")
async def update_cliente(
    cliente_id: str, payload: ClienteUpdate, _: dict = Depends(require_admin)
):
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Sin cambios")
    result = await clientes_col.update_one({"id": cliente_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return await clientes_col.find_one({"id": cliente_id}, {"_id": 0})


@api_router.delete("/admin/clientes/{cliente_id}")
async def delete_cliente(cliente_id: str, _: dict = Depends(require_admin)):
    result = await clientes_col.delete_one({"id": cliente_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    # Cascade: delete sucursales
    await sucursales_col.delete_many({"cliente_id": cliente_id})
    return {"ok": True}


# ----------------- Admin: Sucursales -----------------
@api_router.post("/admin/sucursales", status_code=201)
async def create_sucursal(payload: SucursalCreate, _: dict = Depends(require_admin)):
    cliente = await clientes_col.find_one({"id": payload.cliente_id})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    doc = {
        "id": str(uuid.uuid4()),
        **payload.model_dump(),
        "created_at": now_iso(),
    }
    await sucursales_col.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api_router.get("/admin/sucursales")
async def list_sucursales(
    cliente_id: Optional[str] = Query(None), _: dict = Depends(require_admin)
):
    q = {"cliente_id": cliente_id} if cliente_id else {}
    cursor = sucursales_col.find(q, {"_id": 0}).sort("created_at", -1)
    return await cursor.to_list(1000)


@api_router.patch("/admin/sucursales/{sucursal_id}")
async def update_sucursal(
    sucursal_id: str, payload: SucursalUpdate, _: dict = Depends(require_admin)
):
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Sin cambios")
    result = await sucursales_col.update_one(
        {"id": sucursal_id}, {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    return await sucursales_col.find_one({"id": sucursal_id}, {"_id": 0})


@api_router.delete("/admin/sucursales/{sucursal_id}")
async def delete_sucursal(sucursal_id: str, _: dict = Depends(require_admin)):
    result = await sucursales_col.delete_one({"id": sucursal_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    return {"ok": True}


# ----------------- Admin: Órdenes -----------------
async def _send_assignment_whatsapp(orden_id: str) -> dict:
    """Helper: sends WhatsApp to the technician currently assigned to an orden."""
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    if not o or not o.get("tecnico_id"):
        return {"ok": False, "mode": "no_tecnico", "detail": "Sin técnico asignado"}
    tecnico = await users_col.find_one(
        {"id": o["tecnico_id"]}, {"_id": 0, "hashed_password": 0}
    )
    if not tecnico:
        return {"ok": False, "mode": "no_tecnico", "detail": "Técnico no encontrado"}
    cliente = await clientes_col.find_one({"id": o["cliente_id"]}, {"_id": 0}) or {}
    sucursal = await sucursales_col.find_one({"id": o["sucursal_id"]}, {"_id": 0}) or {}
    msg = build_assignment_message(
        tecnico_nombre=tecnico.get("nombre", ""),
        numero=o.get("numero", ""),
        cliente=cliente.get("nombre_fantasia") or cliente.get("nombre", "—"),
        comercio=sucursal.get("nombre", "—"),
        codigo_comercio=sucursal.get("codigo_comercio", "—"),
        direccion=sucursal.get("direccion", "—"),
        prioridad=o.get("prioridad", "media"),
        fecha_limite=o.get("fecha_limite"),
        pin_pads=o.get("pin_pads") or [],
        rut=cliente.get("rut"),
        razon_social=cliente.get("nombre"),
        nombre_fantasia=cliente.get("nombre_fantasia"),
        comuna=sucursal.get("comuna"),
        region=sucursal.get("region"),
        fecha_ejecucion=o.get("fecha_ejecucion"),
    )
    result = await send_whatsapp(tecnico.get("telefono"), msg)
    # store last notification result on the order
    await ordenes_col.update_one(
        {"id": orden_id},
        {"$set": {"whatsapp_last": {**result, "to": tecnico.get("telefono"), "at": now_iso()}}},
    )
    return result


@api_router.post("/admin/ordenes", status_code=201)
async def create_orden(payload: OrdenCreate, admin: dict = Depends(require_admin)):
    # validations
    cliente = await clientes_col.find_one({"id": payload.cliente_id})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    sucursal = await sucursales_col.find_one(
        {"id": payload.sucursal_id, "cliente_id": payload.cliente_id}
    )
    if not sucursal:
        raise HTTPException(
            status_code=404, detail="Comercio no encontrado para este cliente"
        )
    if payload.tecnico_id:
        tecnico = await users_col.find_one(
            {"id": payload.tecnico_id, "role": "tecnico"}
        )
        if not tecnico:
            raise HTTPException(status_code=404, detail="Técnico no encontrado")

    # Generate numero correlativo
    count = await ordenes_col.count_documents({})
    numero = f"OS-{datetime.now().year}-{count + 1:04d}"

    # Auto-load pin_pads from prior orders of this comercio (each DDLL=1 pin pad)
    # so the technician sees all pin pads to attend, not an empty list.
    pin_pads_inherit: list = []
    seen_keys: set = set()
    prior_ordenes = await ordenes_col.find(
        {"sucursal_id": payload.sucursal_id}, {"_id": 0, "pin_pads": 1, "serie": 1, "modelo": 1, "ddll": 1}
    ).to_list(2000)
    for po in prior_ordenes:
        for pp in po.get("pin_pads") or []:
            key = (pp.get("serie") or "") + "|" + (pp.get("ddll") or "")
            if not key.strip("|") or key in seen_keys:
                continue
            seen_keys.add(key)
            pin_pads_inherit.append(
                {
                    "id": str(uuid.uuid4()),
                    "serie": pp.get("serie"),
                    "modelo": pp.get("modelo"),
                    "ddll": pp.get("ddll"),
                    "fotos": [],
                    "estado": "pendiente",
                }
            )
        # legacy single-pinpad ordenes (no pin_pads array)
        if not po.get("pin_pads") and (po.get("serie") or po.get("ddll")):
            key = (po.get("serie") or "") + "|" + (po.get("ddll") or "")
            if key.strip("|") and key not in seen_keys:
                seen_keys.add(key)
                pin_pads_inherit.append(
                    {
                        "id": str(uuid.uuid4()),
                        "serie": po.get("serie"),
                        "modelo": po.get("modelo"),
                        "ddll": po.get("ddll"),
                        "fotos": [],
                        "estado": "pendiente",
                    }
                )

    # If admin provided a single serie/ddll on the form, ensure it is in pin_pads
    if payload.serie or payload.ddll:
        key = (payload.serie or "") + "|" + (payload.ddll or "")
        if key.strip("|") and key not in seen_keys:
            pin_pads_inherit.append(
                {
                    "id": str(uuid.uuid4()),
                    "serie": payload.serie,
                    "modelo": payload.modelo,
                    "ddll": payload.ddll,
                    "fotos": [],
                    "estado": "pendiente",
                }
            )

    doc = {
        "id": str(uuid.uuid4()),
        "numero": numero,
        "cliente_id": payload.cliente_id,
        "sucursal_id": payload.sucursal_id,
        "tecnico_id": payload.tecnico_id,
        "titulo": payload.titulo,
        "descripcion": payload.descripcion,
        "prioridad": payload.prioridad,
        "serie": payload.serie,
        "modelo": payload.modelo,
        "ddll": payload.ddll,
        "pin_pads": pin_pads_inherit,
        "fecha_limite": payload.fecha_limite,
        "estado": "pendiente",
        "evidencia_base64": None,
        "notas_tecnico": None,
        "created_by": admin["id"],
        "created_at": now_iso(),
        "started_at": None,
        "finalized_at": None,
        "whatsapp_last": None,
    }
    await ordenes_col.insert_one(doc)

    if payload.tecnico_id:
        await _send_assignment_whatsapp(doc["id"])

    o = await ordenes_col.find_one({"id": doc["id"]}, {"_id": 0})
    return await enrich_orden(o)


@api_router.patch("/admin/ordenes/{orden_id}/asignar")
async def asignar_tecnico(
    orden_id: str, payload: OrdenAsignar, _: dict = Depends(require_admin)
):
    orden = await ordenes_col.find_one({"id": orden_id})
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    tecnico = await users_col.find_one(
        {"id": payload.tecnico_id, "role": "tecnico"}
    )
    if not tecnico:
        raise HTTPException(status_code=404, detail="Técnico no encontrado")
    await ordenes_col.update_one(
        {"id": orden_id}, {"$set": {"tecnico_id": payload.tecnico_id}}
    )
    wa = await _send_assignment_whatsapp(orden_id)
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    enriched = await enrich_orden(o)
    return {"orden": enriched, "whatsapp": wa}


@api_router.post("/admin/ordenes/{orden_id}/reenviar-whatsapp")
async def reenviar_whatsapp(orden_id: str, _: dict = Depends(require_admin)):
    """Re-send the WhatsApp assignment message to the technician currently assigned.

    Use this when the previous send failed, the technician didn't receive it, or
    the admin simply wants to resend the order details.
    """
    orden = await ordenes_col.find_one({"id": orden_id})
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if not orden.get("tecnico_id"):
        raise HTTPException(
            status_code=400,
            detail="La orden no tiene técnico asignado. Asigna un técnico primero.",
        )
    wa = await _send_assignment_whatsapp(orden_id)
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    enriched = await enrich_orden(o)
    return {"orden": enriched, "whatsapp": wa}



# ----------------- Settings (config) -----------------

settings_col = db.app_settings

DEFAULT_SETTINGS = {
    "auto_asignacion_masiva": False,
    "max_ordenes_tecnico_dia": 25,
    "minutos_por_pinpad": 10,
}


async def _get_settings() -> dict:
    s = await settings_col.find_one({"_id": "main"}, {"_id": 0})
    if not s:
        await settings_col.insert_one({"_id": "main", **DEFAULT_SETTINGS})
        return DEFAULT_SETTINGS.copy()
    return {**DEFAULT_SETTINGS, **s}


@api_router.get("/admin/settings")
async def get_settings(_: dict = Depends(require_admin)):
    return await _get_settings()


class SettingsUpdate(BaseModel):
    auto_asignacion_masiva: Optional[bool] = None
    max_ordenes_tecnico_dia: Optional[int] = None
    minutos_por_pinpad: Optional[int] = None


@api_router.put("/admin/settings")
async def update_settings(
    payload: SettingsUpdate, _: dict = Depends(require_admin)
):
    upd = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not upd:
        return await _get_settings()
    await settings_col.update_one(
        {"_id": "main"}, {"$set": upd}, upsert=True
    )
    return await _get_settings()


async def _auto_assign_if_enabled() -> dict:
    """Si la auto-asignación está ON, distribuye órdenes pendientes sin
    técnico a técnicos cercanos. Best-effort: nunca falla la operación
    que la invoca."""
    try:
        s = await _get_settings()
        if not s.get("auto_asignacion_masiva"):
            return {"enabled": False, "asignadas": 0}
        # Reusa la lógica de asignación masiva
        max_t = s.get("max_ordenes_tecnico_dia", 25)
        tecnicos = await users_col.find(
            {"role": "tecnico"}, {"_id": 0, "hashed_password": 0}
        ).to_list(500)
        if not tecnicos:
            return {"enabled": True, "asignadas": 0, "motivo": "sin_tecnicos"}
        ordenes = await ordenes_col.find(
            {"estado": "pendiente", "tecnico_id": {"$in": [None, ""]}},
            {"_id": 0},
        ).to_list(2000)
        if not ordenes:
            return {"enabled": True, "asignadas": 0}
        sucursal_ids = list({o.get("sucursal_id") for o in ordenes if o.get("sucursal_id")})
        sucs_map = {}
        if sucursal_ids:
            sucs = await sucursales_col.find(
                {"id": {"$in": sucursal_ids}}, {"_id": 0}
            ).to_list(2000)
            sucs_map = {x["id"]: x for x in sucs}
        carga = {}
        for t in tecnicos:
            carga[t["id"]] = await ordenes_col.count_documents(
                {"tecnico_id": t["id"], "estado": {"$in": ["pendiente", "en_progreso"]}}
            )
        asignadas = 0
        for o in ordenes:
            suc = sucs_map.get(o.get("sucursal_id"), {})
            comuna_orden = (suc.get("comuna") or "").strip()
            cands = sorted(
                tecnicos,
                key=lambda t: (
                    _comuna_match_score(t.get("comuna"), comuna_orden),
                    carga.get(t["id"], 0),
                ),
            )
            elegido = next((t for t in cands if carga.get(t["id"], 0) < max_t), None)
            if not elegido:
                continue
            await ordenes_col.update_one(
                {"id": o["id"]}, {"$set": {"tecnico_id": elegido["id"]}}
            )
            carga[elegido["id"]] = carga.get(elegido["id"], 0) + 1
            asignadas += 1
            try:
                await _send_assignment_whatsapp(o["id"])
            except Exception:
                pass
        return {"enabled": True, "asignadas": asignadas}
    except Exception as e:
        logger.warning("[auto_assign] error: %s", e)
        return {"enabled": False, "error": str(e)}


@api_router.post("/admin/ordenes/upload-excel")
async def upload_ordenes_excel(
    file: UploadFile = File(...),
    fecha_limite: Optional[str] = Form(None),
    prioridad: str = Form("media"),
    admin: dict = Depends(require_admin),
):
    """Upload Excel file (BASE FEMSA REGIONES style) to bulk-create ordenes.

    Strategy:
      - Sheet "FEMSA" (if present): master inventory of pin pads
        Columns: RUT, NOM_RAZON_SOCIAL, NOM_FANTASIA, CC, DIRECCION, COMUNA,
                 #REGION, M_MODELO, SERIEPI, DDLL
      - Sheet "CC" (if present): work orders to create (one orden per row).
        Columns include: RUT, NOM_RAZON_SOCIAL, NOM_FANTASIA, CC, DIRECCION,
                         COMUNA, #REGION, Cantidad, Fecha
      - For each CC row: locate matching pin pads from FEMSA inventory by
        (rut, CC); create ONE orden with that list of pin_pads.
      - If only FEMSA exists: group its rows by (rut, CC) and create one
        orden per group with all its pin pads.
    """
    if prioridad not in ("baja", "media", "alta"):
        prioridad = "media"
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel inválido: {e}")

    sheet_femsa = None
    sheet_cc = None
    for s in wb.sheetnames:
        low = s.strip().lower()
        if low == "femsa":
            sheet_femsa = s
        elif low == "cc":
            sheet_cc = s
    if not sheet_femsa and not sheet_cc:
        # fallback: first sheet
        sheet_femsa = wb.sheetnames[0]

    def make_cell_fn(headers_row):
        h = [str(h).strip() if h is not None else "" for h in headers_row]
        idx = {x.upper(): i for i, x in enumerate(h)}

        def get(row, name):
            i = idx.get(name.upper())
            if i is None or i >= len(row):
                return None
            v = row[i]
            if v is None:
                return None
            if hasattr(v, "isoformat"):
                # date/datetime
                try:
                    return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
                except Exception:
                    return str(v)
            s = str(v).strip()
            return s if s else None

        return get

    # Build FEMSA inventory: (rut_or_razon, cc) -> [pin_pad_dict]
    inventory: dict = {}
    if sheet_femsa:
        ws = wb[sheet_femsa]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            get = make_cell_fn(rows[0])
            for row in rows[1:]:
                if not row or all(c is None for c in row):
                    continue
                rut = get(row, "RUT")
                razon = get(row, "NOM_RAZON_SOCIAL")
                cc = get(row, "CC")
                if not cc:
                    continue
                modelo = get(row, "M_MODELO") or get(row, "MODELO")
                serie = get(row, "SERIEPI") or get(row, "SERIE")
                ddll = get(row, "DDLL")
                if not (serie or ddll):
                    continue
                key = ((rut or razon or "").strip(), cc.strip())
                inventory.setdefault(key, []).append(
                    {
                        "id": str(uuid.uuid4()),
                        "serie": serie,
                        "modelo": modelo,
                        "ddll": ddll,
                        "completed": False,
                        "evidencia_base64": None,
                        "notas": None,
                        "completed_at": None,
                    }
                )

    summary = {
        "total_rows": 0,
        "ordenes_creadas": 0,
        "clientes_creados": 0,
        "comercios_creados": 0,
        "pin_pads_total": 0,
        "errores": [],
    }

    async def upsert_cliente(rut, razon, fantasia):
        cli_query = {"rut": rut} if rut else {"nombre": razon}
        cli = await clientes_col.find_one(cli_query)
        if not cli:
            cli = {
                "id": str(uuid.uuid4()),
                "nombre": razon or fantasia,
                "nombre_fantasia": fantasia,
                "rut": rut,
                "contacto": None,
                "email": None,
                "telefono": None,
                "direccion": None,
                "created_at": now_iso(),
            }
            await clientes_col.insert_one(cli)
            summary["clientes_creados"] += 1
        else:
            if fantasia and not cli.get("nombre_fantasia"):
                await clientes_col.update_one(
                    {"id": cli["id"]}, {"$set": {"nombre_fantasia": fantasia}}
                )
        return cli

    async def upsert_comercio(cliente, cc, direccion, comuna, region, fantasia, razon):
        suc = await sucursales_col.find_one(
            {"cliente_id": cliente["id"], "codigo_comercio": cc}
        )
        if not suc:
            # Intentar geocodificar (best-effort, no bloqueante)
            lat = None
            lng = None
            if direccion:
                try:
                    r = await geocode_address(db, direccion, comuna, region)
                    if r:
                        lat, lng, _ = r
                except Exception as e:
                    logger.warning("[geocode sucursal] err: %s", e)

            suc = {
                "id": str(uuid.uuid4()),
                "cliente_id": cliente["id"],
                "nombre": (fantasia or razon or "") + f" - {cc}",
                "codigo_comercio": cc,
                "direccion": direccion,
                "comuna": comuna,
                "region": region,
                "lat": lat,
                "lng": lng,
                "telefono": None,
                "encargado": None,
                "created_at": now_iso(),
            }
            await sucursales_col.insert_one(suc)
            summary["comercios_creados"] += 1
        elif (suc.get("lat") is None or suc.get("lng") is None) and (suc.get("direccion") or direccion):
            # Sucursal existe pero sin coords → intentar geocodificar
            try:
                r = await geocode_address(
                    db,
                    suc.get("direccion") or direccion,
                    suc.get("comuna") or comuna,
                    suc.get("region") or region,
                )
                if r:
                    suc["lat"], suc["lng"], _ = r
                    await sucursales_col.update_one(
                        {"id": suc["id"]},
                        {"$set": {"lat": suc["lat"], "lng": suc["lng"]}},
                    )
            except Exception as e:
                logger.warning("[geocode sucursal existing] err: %s", e)
        return suc

    async def create_orden(cliente, suc, cc, fecha_lim, pin_pads, direccion):
        # fresh copies of pin_pads (new ids each time)
        fresh_pp = [
            {
                **pp,
                "id": str(uuid.uuid4()),
                "completed": False,
                "evidencia_base64": None,
                "notas": None,
                "completed_at": None,
            }
            for pp in pin_pads
        ]
        count = await ordenes_col.count_documents({})
        numero = f"OS-{datetime.now().year}-{count + 1:04d}"
        num_pp = len(fresh_pp)
        titulo = f"Actualización Pin Pads · CC {cc}"
        descripcion = (
            f"Actualizar {num_pp} pin pad{'s' if num_pp != 1 else ''} en el "
            f"comercio CC {cc} ({direccion or '—'})."
        )
        doc = {
            "id": str(uuid.uuid4()),
            "numero": numero,
            "cliente_id": cliente["id"],
            "sucursal_id": suc["id"],
            "tecnico_id": None,
            "titulo": titulo,
            "descripcion": descripcion,
            "prioridad": prioridad,
            "serie": None,
            "modelo": None,
            "ddll": None,
            "fecha_limite": fecha_lim,
            "estado": "pendiente",
            "evidencia_base64": None,
            "notas_tecnico": None,
            "pin_pads": fresh_pp,
            "created_by": admin["id"],
            "created_at": now_iso(),
            "started_at": None,
            "finalized_at": None,
            "whatsapp_last": None,
        }
        await ordenes_col.insert_one(doc)
        summary["ordenes_creadas"] += 1
        summary["pin_pads_total"] += num_pp

    if sheet_cc:
        ws = wb[sheet_cc]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            get = make_cell_fn(rows[0])
            for row in rows[1:]:
                if not row or all(c is None for c in row):
                    continue
                summary["total_rows"] += 1
                try:
                    rut = get(row, "RUT")
                    razon = get(row, "NOM_RAZON_SOCIAL")
                    fantasia = get(row, "NOM_FANTASIA")
                    cc = get(row, "CC")
                    direccion = get(row, "DIRECCION")
                    comuna = get(row, "COMUNA")
                    region = get(row, "#REGION") or get(row, "REGION")
                    fecha_row = get(row, "Fecha") or get(row, "FECHA")
                    if not cc:
                        summary["errores"].append("Fila sin CC")
                        continue
                    fecha_lim = fecha_row or fecha_limite

                    cliente = await upsert_cliente(rut, razon, fantasia)
                    suc = await upsert_comercio(
                        cliente, cc, direccion, comuna, region, fantasia, razon
                    )
                    # look up pin pads from inventory
                    key = ((rut or razon or "").strip(), cc.strip())
                    pin_pads = inventory.get(key, [])
                    await create_orden(cliente, suc, cc, fecha_lim, pin_pads, direccion)
                except Exception as e:
                    summary["errores"].append(f"CC {cc}: {str(e)[:120]}")
    else:
        # Only FEMSA exists: group by (rut, cc)
        groups: dict = {}
        ws = wb[sheet_femsa]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            get = make_cell_fn(rows[0])
            for row in rows[1:]:
                if not row or all(c is None for c in row):
                    continue
                summary["total_rows"] += 1
                rut = get(row, "RUT")
                razon = get(row, "NOM_RAZON_SOCIAL")
                fantasia = get(row, "NOM_FANTASIA")
                cc = get(row, "CC")
                if not cc:
                    continue
                direccion = get(row, "DIRECCION")
                comuna = get(row, "COMUNA")
                region = get(row, "#REGION") or get(row, "REGION")
                modelo = get(row, "M_MODELO")
                serie = get(row, "SERIEPI")
                ddll = get(row, "DDLL")
                gkey = ((rut or razon or "").strip(), cc.strip())
                g = groups.setdefault(
                    gkey,
                    {
                        "rut": rut,
                        "razon": razon,
                        "fantasia": fantasia,
                        "cc": cc,
                        "direccion": direccion,
                        "comuna": comuna,
                        "region": region,
                        "pin_pads": [],
                    },
                )
                if serie or ddll:
                    g["pin_pads"].append(
                        {
                            "id": str(uuid.uuid4()),
                            "serie": serie,
                            "modelo": modelo,
                            "ddll": ddll,
                            "completed": False,
                            "evidencia_base64": None,
                            "notas": None,
                            "completed_at": None,
                        }
                    )

        for _, g in groups.items():
            try:
                cliente = await upsert_cliente(g["rut"], g["razon"], g["fantasia"])
                suc = await upsert_comercio(
                    cliente,
                    g["cc"],
                    g["direccion"],
                    g["comuna"],
                    g["region"],
                    g["fantasia"],
                    g["razon"],
                )
                await create_orden(
                    cliente, suc, g["cc"], fecha_limite, g["pin_pads"], g["direccion"]
                )
            except Exception as e:
                summary["errores"].append(f"CC {g['cc']}: {str(e)[:120]}")

    # Auto-asignación masiva si está habilitada
    auto = await _auto_assign_if_enabled()
    summary["auto_asignacion"] = auto
    return summary


@api_router.post("/admin/ordenes/cleanup")
async def cleanup_ordenes(
    days: int = Query(60, ge=1, le=365), _: dict = Depends(require_admin)
):
    """Elimina órdenes con created_at de hace más de N días (default 60)."""
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    cutoff_iso = cutoff.isoformat()
    r = await ordenes_col.delete_many({"created_at": {"$lt": cutoff_iso}})
    logger.info(f"[Cleanup] {r.deleted_count} órdenes eliminadas (> {days} días)")
    return {"deleted": r.deleted_count, "days": days, "cutoff": cutoff_iso}


@api_router.post("/admin/ordenes/reset")
async def reset_ordenes(_: dict = Depends(require_admin)):
    """Elimina TODAS las órdenes (mantiene clientes, comercios y técnicos)."""
    r = await ordenes_col.delete_many({})
    return {"deleted": r.deleted_count}


@api_router.get("/admin/comercios/export.csv")
async def export_comercios_csv(_: dict = Depends(require_admin)):
    """Exporta CSV con: Cliente, CC, Dirección, Comuna, Región, Serie, DDLL, Modelo."""
    sucursales = await sucursales_col.find({}, {"_id": 0}).to_list(5000)
    lines = ["Cliente,CC,Direccion,Comuna,Region,Serie,DDLL,Modelo"]

    def csv_escape(v):
        if v is None:
            return ""
        s = str(v).replace('"', '""')
        if "," in s or "\n" in s or '"' in s:
            return f'"{s}"'
        return s

    for s in sucursales:
        cliente = await clientes_col.find_one({"id": s.get("cliente_id")}, {"_id": 0})
        cli_name = (cliente or {}).get("nombre_fantasia") or (cliente or {}).get("nombre", "")
        ordenes = await ordenes_col.find(
            {"sucursal_id": s["id"]}, {"_id": 0}
        ).to_list(2000)
        seen = set()
        pp_list = []
        for o in ordenes:
            for pp in o.get("pin_pads") or []:
                k = (pp.get("serie") or "") + "|" + (pp.get("ddll") or "")
                if k not in seen:
                    seen.add(k)
                    pp_list.append(pp)
        if not pp_list:
            lines.append(
                ",".join(
                    csv_escape(x)
                    for x in [
                        cli_name,
                        s.get("codigo_comercio"),
                        s.get("direccion"),
                        s.get("comuna"),
                        s.get("region"),
                        "",
                        "",
                        "",
                    ]
                )
            )
        for pp in pp_list:
            lines.append(
                ",".join(
                    csv_escape(x)
                    for x in [
                        cli_name,
                        s.get("codigo_comercio"),
                        s.get("direccion"),
                        s.get("comuna"),
                        s.get("region"),
                        pp.get("serie"),
                        pp.get("ddll"),
                        pp.get("modelo"),
                    ]
                )
            )
    csv_text = "\n".join(lines) + "\n"
    return StreamingResponse(
        io.BytesIO(csv_text.encode("utf-8-sig")),
        media_type="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="comercios_mvg.csv"'
        },
    )


@api_router.get("/admin/comercios")
async def list_comercios(_: dict = Depends(require_admin)):
    """Returns all comercios with cliente info and pin_pads summary derived from ordenes."""
    sucursales = await sucursales_col.find({}, {"_id": 0}).to_list(5000)
    out = []
    for s in sucursales:
        cliente = await clientes_col.find_one(
            {"id": s.get("cliente_id")}, {"_id": 0}
        )
        # Aggregate pin_pads across all ordenes of this comercio
        ordenes = await ordenes_col.find(
            {"sucursal_id": s["id"]}, {"_id": 0}
        ).to_list(2000)
        pin_pads_map = {}
        ordenes_count = 0
        finalizadas_count = 0
        for o in ordenes:
            ordenes_count += 1
            if o.get("estado") == "finalizada":
                finalizadas_count += 1
            for pp in o.get("pin_pads") or []:
                key = (pp.get("serie") or "") + "|" + (pp.get("ddll") or "")
                if key not in pin_pads_map:
                    pin_pads_map[key] = {
                        "serie": pp.get("serie"),
                        "modelo": pp.get("modelo"),
                        "ddll": pp.get("ddll"),
                    }
            # legacy single-pinpad ordenes
            if not o.get("pin_pads") and (o.get("serie") or o.get("ddll")):
                key = (o.get("serie") or "") + "|" + (o.get("ddll") or "")
                if key not in pin_pads_map:
                    pin_pads_map[key] = {
                        "serie": o.get("serie"),
                        "modelo": o.get("modelo"),
                        "ddll": o.get("ddll"),
                    }
        out.append(
            {
                **s,
                "cliente": cliente,
                "pin_pads": list(pin_pads_map.values()),
                "pin_pads_count": len(pin_pads_map),
                "ordenes_count": ordenes_count,
                "ordenes_finalizadas": finalizadas_count,
            }
        )
    # sort by cliente name then cc
    out.sort(
        key=lambda x: (
            (x.get("cliente") or {}).get("nombre_fantasia")
            or (x.get("cliente") or {}).get("nombre", ""),
            x.get("codigo_comercio") or "",
        )
    )
    return out


@api_router.get("/admin/ordenes/{orden_id}/pdf")
async def orden_pdf(orden_id: str, _: dict = Depends(require_admin)):
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    cliente = await clientes_col.find_one({"id": o.get("cliente_id")}, {"_id": 0}) or {}
    sucursal = await sucursales_col.find_one({"id": o.get("sucursal_id")}, {"_id": 0}) or {}
    tecnico = await users_col.find_one(
        {"id": o.get("tecnico_id")}, {"_id": 0, "hashed_password": 0}
    ) or {}
    pdf_bytes = build_orden_pdf(o, cliente, sucursal, tecnico)
    filename = f"orden_{o.get('numero', orden_id)}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/admin/ordenes")
async def list_ordenes_admin(
    estado: Optional[str] = Query(None),
    prioridad: Optional[str] = Query(None),
    tecnico_id: Optional[str] = Query(None),
    _: dict = Depends(require_admin),
):
    q = {}
    if estado:
        q["estado"] = estado
    if prioridad:
        q["prioridad"] = prioridad
    if tecnico_id:
        q["tecnico_id"] = tecnico_id
    cursor = ordenes_col.find(q, {"_id": 0}).sort("created_at", -1)
    items = await cursor.to_list(1000)
    return [await enrich_orden(o) for o in items]


@api_router.get("/admin/stats")
async def admin_stats(_: dict = Depends(require_admin)):
    total = await ordenes_col.count_documents({})
    en_progreso = await ordenes_col.count_documents({"estado": "en_progreso"})
    pendientes = await ordenes_col.count_documents({"estado": "pendiente"})
    finalizadas = await ordenes_col.count_documents({"estado": "finalizada"})
    reagendadas = await ordenes_col.count_documents({"estado": "reagendada"})
    total_clientes = await clientes_col.count_documents({})
    total_tecnicos = await users_col.count_documents({"role": "tecnico"})

    # Calcular Pin Pads pendientes y técnicos necesarios (25 pinpads/téc/día)
    PINPADS_POR_TECNICO_DIA = 25
    pendientes_docs = await ordenes_col.find(
        {"estado": {"$in": ["pendiente", "en_progreso"]}},
        {"_id": 0, "pin_pads": 1},
    ).to_list(5000)
    pin_pads_pendientes = 0
    for o in pendientes_docs:
        for pp in (o.get("pin_pads") or []):
            if not pp.get("completed"):
                pin_pads_pendientes += 1
    import math as _math
    tecnicos_necesarios = _math.ceil(pin_pads_pendientes / PINPADS_POR_TECNICO_DIA) if pin_pads_pendientes else 0
    horas_estimadas = round(pin_pads_pendientes * 10 / 60, 1)

    return {
        "total_ordenes": total,
        "en_progreso": en_progreso,
        "pendientes": pendientes,
        "finalizadas": finalizadas,
        "reagendadas": reagendadas,
        "total_clientes": total_clientes,
        "total_tecnicos": total_tecnicos,
        "pin_pads_pendientes": pin_pads_pendientes,
        "tecnicos_necesarios": tecnicos_necesarios,
        "horas_totales_estimadas": horas_estimadas,
        "pin_pads_por_tecnico_dia": PINPADS_POR_TECNICO_DIA,
    }


@api_router.patch("/admin/ordenes/{orden_id}")
async def update_orden(
    orden_id: str, payload: OrdenUpdate, _: dict = Depends(require_admin)
):
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Sin cambios")
    result = await ordenes_col.update_one({"id": orden_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


@api_router.delete("/admin/ordenes/{orden_id}")
async def delete_orden(orden_id: str, _: dict = Depends(require_admin)):
    result = await ordenes_col.delete_one({"id": orden_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    return {"ok": True}


# ----------------- Shared orden detail -----------------
@api_router.get("/ordenes/{orden_id}")
async def get_orden(orden_id: str, current: dict = Depends(get_current_user)):
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    # Scope check for tecnico
    if current["role"] == "tecnico" and o.get("tecnico_id") != current["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    # Lazy backfill: if pin_pads is empty, try to inherit from previous orders
    # of the same comercio (sucursal). Each DDLL/Serie is one pin pad.
    if not o.get("pin_pads"):
        seen_keys: set = set()
        pp_inherit: list = []
        prior = await ordenes_col.find(
            {
                "sucursal_id": o.get("sucursal_id"),
                "id": {"$ne": o["id"]},
            },
            {"_id": 0, "pin_pads": 1, "serie": 1, "modelo": 1, "ddll": 1},
        ).to_list(2000)
        for po in prior:
            for pp in po.get("pin_pads") or []:
                key = (pp.get("serie") or "") + "|" + (pp.get("ddll") or "")
                if not key.strip("|") or key in seen_keys:
                    continue
                seen_keys.add(key)
                pp_inherit.append(
                    {
                        "id": str(uuid.uuid4()),
                        "serie": pp.get("serie"),
                        "modelo": pp.get("modelo"),
                        "ddll": pp.get("ddll"),
                        "fotos": [],
                        "estado": "pendiente",
                    }
                )
            if not po.get("pin_pads") and (po.get("serie") or po.get("ddll")):
                key = (po.get("serie") or "") + "|" + (po.get("ddll") or "")
                if key.strip("|") and key not in seen_keys:
                    seen_keys.add(key)
                    pp_inherit.append(
                        {
                            "id": str(uuid.uuid4()),
                            "serie": po.get("serie"),
                            "modelo": po.get("modelo"),
                            "ddll": po.get("ddll"),
                            "fotos": [],
                            "estado": "pendiente",
                        }
                    )
        if pp_inherit:
            await ordenes_col.update_one(
                {"id": o["id"]}, {"$set": {"pin_pads": pp_inherit}}
            )
            o["pin_pads"] = pp_inherit
    return await enrich_orden(o)


# ----------------- Tecnico routes -----------------
@api_router.get("/tecnico/ordenes")
async def list_ordenes_tecnico(
    estado: Optional[str] = Query(None), tec: dict = Depends(require_tecnico)
):
    q = {"tecnico_id": tec["id"]}
    if estado:
        q["estado"] = estado
    cursor = ordenes_col.find(q, {"_id": 0}).sort("created_at", -1)
    items = await cursor.to_list(1000)
    return [await enrich_orden(o) for o in items]


@api_router.get("/tecnico/stats")
async def tecnico_stats(tec: dict = Depends(require_tecnico)):
    base = {"tecnico_id": tec["id"]}
    total = await ordenes_col.count_documents(base)
    pendientes = await ordenes_col.count_documents({**base, "estado": "pendiente"})
    en_progreso = await ordenes_col.count_documents(
        {**base, "estado": "en_progreso"}
    )
    finalizadas = await ordenes_col.count_documents(
        {**base, "estado": "finalizada"}
    )
    return {
        "total": total,
        "pendientes": pendientes,
        "en_progreso": en_progreso,
        "finalizadas": finalizadas,
    }


@api_router.patch("/tecnico/ordenes/{orden_id}/iniciar")
async def iniciar_orden(orden_id: str, tec: dict = Depends(require_tecnico)):
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    if o.get("estado") != "pendiente":
        raise HTTPException(status_code=400, detail="La orden no está pendiente")
    await ordenes_col.update_one(
        {"id": orden_id},
        {"$set": {"estado": "en_progreso", "started_at": now_iso()}},
    )
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


@api_router.patch("/tecnico/ordenes/{orden_id}/pinpad/{pinpad_id}")
async def update_pinpad(
    orden_id: str,
    pinpad_id: str,
    payload: PinPadUpdate,
    tec: dict = Depends(require_tecnico),
):
    """Técnico marca UN pin pad como actualizado con evidencia fotográfica.

    Requirements:
    - evidencia_base64 (foto)
    - materiales_usados (>=1 ítem) OR sin_suministros=True
    """
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")

    # ---- Validar fotos (4 fotos del protocolo MVG con backward compat) ----
    # Nuevo flujo: foto_antes + foto_despues OBLIGATORIAS, las otras 2 opcionales.
    # Legacy: si solo viene evidencia_base64, lo aceptamos como antes Y después
    # (no podemos partirlo, así que lo guardamos como evidencia "principal").
    has_new_required = bool(payload.foto_antes_base64) and bool(payload.foto_despues_base64)
    has_legacy = bool(payload.evidencia_base64)
    if not has_new_required and not has_legacy:
        raise HTTPException(
            status_code=400,
            detail=(
                "Se requieren al menos 2 fotos: Pinpad antes de actualizar y "
                "Pinpad después de actualizar."
            ),
        )

    # Compress incoming photos
    compressed_antes = compress_evidencia(payload.foto_antes_base64) if payload.foto_antes_base64 else None
    compressed_dm = compress_evidencia(payload.foto_descarga_master_base64) if payload.foto_descarga_master_base64 else None
    compressed_despues = compress_evidencia(payload.foto_despues_base64) if payload.foto_despues_base64 else None
    compressed_cv = compress_evidencia(payload.foto_comprobante_venta_base64) if payload.foto_comprobante_venta_base64 else None
    # Foto principal de compat: usa "despues" si existe, sino el legacy
    compressed_evidence = (
        compressed_despues
        or (compress_evidencia(payload.evidencia_base64) if payload.evidencia_base64 else None)
    )

    materiales = payload.materiales_usados or []
    if not materiales and not payload.sin_suministros:
        raise HTTPException(
            status_code=400,
            detail=(
                "Selecciona los materiales utilizados o marca "
                "'No se utilizaron suministros'."
            ),
        )

    # ---- Deduct stock per material (does not block submission on shortage) ----
    consumo_results: List[dict] = []
    if materiales:
        try:
            inv_col = db.inventario_tecnico
            prod_col = db.productos
            consumos_log = db.consumos_materiales
            for item in materiales:
                sku = (item.sku or "").strip()
                if not sku or item.cantidad <= 0:
                    continue
                inv = await inv_col.find_one(
                    {"tecnico_id": tec["id"], "sku": sku}
                )
                prod = await prod_col.find_one({"sku": sku})
                desc = (prod or {}).get("descripcion") or item.descripcion
                current_qty = inv["cantidad"] if inv else 0
                new_qty = current_qty - item.cantidad
                negativo = new_qty < 0
                if inv:
                    await inv_col.update_one(
                        {"id": inv["id"]},
                        {"$set": {"cantidad": new_qty, "updated_at": now_iso()}},
                    )
                else:
                    await inv_col.insert_one({
                        "id": str(uuid.uuid4()),
                        "tecnico_id": tec["id"],
                        "sku": sku,
                        "descripcion": desc,
                        "cantidad": new_qty,
                        "created_at": now_iso(),
                        "updated_at": now_iso(),
                    })
                consumo_results.append({
                    "sku": sku,
                    "descripcion": desc,
                    "cantidad": item.cantidad,
                    "nuevo_stock": new_qty,
                    "negativo": negativo,
                })
            # Audit log
            await consumos_log.insert_one({
                "id": str(uuid.uuid4()),
                "tecnico_id": tec["id"],
                "orden_id": orden_id,
                "pin_pad_id": pinpad_id,
                "items": [it.model_dump() for it in materiales],
                "results": consumo_results,
                "fecha": now_iso(),
            })
        except Exception as e:
            logger.warning("[Consumo pinpad] error: %s", e)

    pin_pads = o.get("pin_pads") or []
    found = False
    all_done = True
    for pp in pin_pads:
        if pp.get("id") == pinpad_id:
            pp["completed"] = True
            # Mantener compat: guarda evidencia_base64 con la foto "después" o legacy
            if compressed_evidence:
                pp["evidencia_base64"] = compressed_evidence
            # Las 4 fotos del protocolo MVG (None si no se enviaron)
            if compressed_antes is not None:
                pp["foto_antes_base64"] = compressed_antes
            if compressed_dm is not None:
                pp["foto_descarga_master_base64"] = compressed_dm
            if compressed_despues is not None:
                pp["foto_despues_base64"] = compressed_despues
            if compressed_cv is not None:
                pp["foto_comprobante_venta_base64"] = compressed_cv
            pp["notas"] = payload.notas
            pp["completed_at"] = now_iso()
            pp["uploaded_at"] = now_iso()  # for 30-min edit window
            pp["materiales_usados"] = [
                {
                    "sku": (item.sku or "").strip(),
                    "descripcion": item.descripcion,
                    "cantidad": item.cantidad,
                }
                for item in materiales
            ]
            pp["sin_suministros"] = bool(payload.sin_suministros) and not materiales
            if payload.lat is not None and payload.lng is not None:
                pp["lat"] = payload.lat
                pp["lng"] = payload.lng
                pp["address"] = payload.address
                pp["accuracy_m"] = payload.accuracy_m
            found = True
        if not pp.get("completed"):
            all_done = False
    if not found:
        raise HTTPException(status_code=404, detail="Pin pad no encontrado")

    set_data = {"pin_pads": pin_pads}
    # auto-start if pendiente
    if o.get("estado") == "pendiente":
        set_data["estado"] = "en_progreso"
        set_data["started_at"] = now_iso()
    # auto-finalize if all done — REQUIRES location to close
    if all_done:
        if payload.lat is None or payload.lng is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Ubicación requerida para finalizar la orden. "
                    "Autoriza la geolocalización en tu navegador."
                ),
            )
        set_data["estado"] = "finalizada"
        set_data["finalized_at"] = now_iso()
        set_data["closed_lat"] = payload.lat
        set_data["closed_lng"] = payload.lng
        set_data["closed_address"] = payload.address
        set_data["closed_accuracy_m"] = payload.accuracy_m

    # Best-effort: if doc would exceed 16MB BSON limit, retro-compress ALL
    # other pin pad photos in this order so the update can succeed.
    try:
        await ordenes_col.update_one({"id": orden_id}, {"$set": set_data})
    except Exception as e:
        if "too large" in str(e).lower() or "DocumentTooLarge" in str(type(e).__name__):
            logger.warning(
                "[update_pinpad] doc too large for orden=%s — recompressing all photos",
                orden_id,
            )
            shrunk = []
            for pp in pin_pads:
                if pp.get("evidencia_base64"):
                    pp["evidencia_base64"] = compress_evidencia(
                        pp["evidencia_base64"], max_side=800, quality=40
                    )
                shrunk.append(pp)
            set_data["pin_pads"] = shrunk
            try:
                await ordenes_col.update_one({"id": orden_id}, {"$set": set_data})
            except Exception as e2:
                logger.error("[update_pinpad] still too large after shrink: %s", e2)
                raise HTTPException(
                    status_code=413,
                    detail=(
                        "Las fotos pesan demasiado para guardarlas en una sola orden. "
                        "Vuelve a tomar la foto con menor calidad o contacta al admin."
                    ),
                )
        else:
            raise
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


@api_router.patch("/tecnico/ordenes/{orden_id}/finalizar")
async def finalizar_orden(
    orden_id: str, payload: OrdenFinalizar, tec: dict = Depends(require_tecnico)
):
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    if o.get("estado") == "finalizada":
        raise HTTPException(status_code=400, detail="Orden ya finalizada")
    if not payload.evidencia_base64:
        raise HTTPException(status_code=400, detail="Evidencia fotográfica requerida")
    await ordenes_col.update_one(
        {"id": orden_id},
        {
            "$set": {
                "estado": "finalizada",
                "evidencia_base64": payload.evidencia_base64,
                "notas_tecnico": payload.notas,
                "finalized_at": now_iso(),
                "closed_lat": payload.lat,
                "closed_lng": payload.lng,
                "closed_address": payload.address,
                "closed_accuracy_m": payload.accuracy_m,
            }
        },
    )
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


# ----- Pin pad photo editing within 30-min window -----
EDIT_WINDOW_MIN = 30


def _can_edit(pp: dict) -> bool:
    ts = pp.get("uploaded_at") or pp.get("completed_at")
    if not ts:
        return False
    try:
        uploaded = datetime.fromisoformat(ts.replace("Z", "+00:00"))
    except Exception:
        return False
    delta = datetime.now(timezone.utc) - uploaded
    return delta < timedelta(minutes=EDIT_WINDOW_MIN)


@api_router.patch("/tecnico/ordenes/{orden_id}/pinpad/{pinpad_id}/foto")
async def replace_pinpad_photo(
    orden_id: str,
    pinpad_id: str,
    payload: PinPadEditPhoto,
    tec: dict = Depends(require_tecnico),
):
    """Replace the photo of a completed pin pad within EDIT_WINDOW_MIN minutes."""
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    pin_pads = o.get("pin_pads") or []
    target = next((p for p in pin_pads if p.get("id") == pinpad_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Pin pad no encontrado")
    if not target.get("completed"):
        raise HTTPException(
            status_code=400, detail="Esta pin pad aún no tiene foto cargada"
        )
    if not _can_edit(target):
        raise HTTPException(
            status_code=403,
            detail=(
                f"La ventana de edición de {EDIT_WINDOW_MIN} minutos ya expiró. "
                "Contacta al admin para correcciones."
            ),
        )
    target["edited_at"] = now_iso()
    # Aplicar las fotos que vengan (cualquier combinación)
    if payload.evidencia_base64:
        target["evidencia_base64"] = compress_evidencia(payload.evidencia_base64)
    if payload.foto_antes_base64:
        target["foto_antes_base64"] = compress_evidencia(payload.foto_antes_base64)
    if payload.foto_descarga_master_base64:
        target["foto_descarga_master_base64"] = compress_evidencia(payload.foto_descarga_master_base64)
    if payload.foto_despues_base64:
        target["foto_despues_base64"] = compress_evidencia(payload.foto_despues_base64)
        # también actualiza evidencia legacy si está
        target["evidencia_base64"] = target["foto_despues_base64"]
    if payload.foto_comprobante_venta_base64:
        target["foto_comprobante_venta_base64"] = compress_evidencia(payload.foto_comprobante_venta_base64)
    if payload.notas is not None:
        target["notas"] = payload.notas
    await ordenes_col.update_one({"id": orden_id}, {"$set": {"pin_pads": pin_pads}})
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


@api_router.delete("/tecnico/ordenes/{orden_id}/pinpad/{pinpad_id}/foto")
async def delete_pinpad_photo(
    orden_id: str,
    pinpad_id: str,
    tec: dict = Depends(require_tecnico),
):
    """Delete the photo (and reset completion) of a pin pad within window."""
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    if o.get("estado") == "finalizada":
        raise HTTPException(
            status_code=400,
            detail=(
                "La orden ya fue finalizada. No se pueden eliminar fotos "
                "después del cierre."
            ),
        )
    pin_pads = o.get("pin_pads") or []
    target = next((p for p in pin_pads if p.get("id") == pinpad_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Pin pad no encontrado")
    if not target.get("completed"):
        raise HTTPException(
            status_code=400, detail="Esta pin pad aún no tiene foto cargada"
        )
    if not _can_edit(target):
        raise HTTPException(
            status_code=403,
            detail=(
                f"La ventana de edición de {EDIT_WINDOW_MIN} minutos ya expiró."
            ),
        )
    target["completed"] = False
    target["evidencia_base64"] = None
    target["notas"] = None
    target["completed_at"] = None
    target["uploaded_at"] = None
    target["lat"] = None
    target["lng"] = None
    target["address"] = None
    target["accuracy_m"] = None
    await ordenes_col.update_one({"id": orden_id}, {"$set": {"pin_pads": pin_pads}})
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


# ============================================================
# NUEVAS FEATURES (Fase 1, 2, 3)
# ============================================================

class PinPadExtraCreate(BaseModel):
    ddll: str
    serie: Optional[str] = None
    modelo: Optional[str] = None
    evidencia_base64: str
    notas: Optional[str] = None
    materiales_usados: Optional[List[MaterialUsadoItem]] = None
    sin_suministros: Optional[bool] = False


@api_router.post("/tecnico/ordenes/{orden_id}/pinpad-extra")
async def add_pinpad_extra(
    orden_id: str,
    payload: PinPadExtraCreate,
    tec: dict = Depends(require_tecnico),
):
    """Técnico agrega un Pin Pad ADICIONAL no listado en la orden.

    Valida que la DDLL tenga la misma longitud y patrón que las existentes,
    sube la foto + materiales y queda visible para el admin.
    """
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    if not payload.ddll or not payload.ddll.strip():
        raise HTTPException(status_code=400, detail="DDLL requerido")

    pin_pads = o.get("pin_pads") or []
    ddll_clean = payload.ddll.strip().upper()

    # Validar formato similar a las DDLL existentes
    existing_ddlls = [pp.get("ddll", "") for pp in pin_pads if pp.get("ddll")]
    if existing_ddlls:
        # Detect length and patrón (alfanumérico)
        ref_len = len(existing_ddlls[0])
        ref_is_alnum = any(c.isalpha() for c in existing_ddlls[0])
        if len(ddll_clean) != ref_len:
            raise HTTPException(
                status_code=400,
                detail=f"Ingresa DDLL correctamente (debe tener {ref_len} caracteres)",
            )
        if ref_is_alnum and not ddll_clean.replace("-", "").isalnum():
            raise HTTPException(
                status_code=400,
                detail="Ingresa DDLL correctamente (solo letras y números)",
            )
        if not ref_is_alnum and not ddll_clean.isdigit():
            raise HTTPException(
                status_code=400,
                detail="Ingresa DDLL correctamente (solo dígitos numéricos)",
            )

    # Verificar duplicado
    if any((pp.get("ddll") or "").strip().upper() == ddll_clean for pp in pin_pads):
        raise HTTPException(
            status_code=400, detail=f"DDLL {ddll_clean} ya existe en esta orden"
        )

    # Validación materiales
    materiales = payload.materiales_usados or []
    if not materiales and not payload.sin_suministros:
        raise HTTPException(
            status_code=400,
            detail='Selecciona materiales o marca "No se utilizaron suministros".',
        )

    # Descuento de stock
    if materiales:
        try:
            inv_col = db.inventario_tecnico
            prod_col = db.productos
            consumos_log = db.consumos_materiales
            for item in materiales:
                if not item.sku or item.cantidad <= 0:
                    continue
                inv = await inv_col.find_one({"tecnico_id": tec["id"], "sku": item.sku})
                prod = await prod_col.find_one({"sku": item.sku})
                desc = (prod or {}).get("descripcion") or item.descripcion
                if inv:
                    await inv_col.update_one(
                        {"id": inv["id"]},
                        {"$set": {"cantidad": inv["cantidad"] - item.cantidad, "updated_at": now_iso()}},
                    )
                else:
                    await inv_col.insert_one({
                        "id": str(uuid.uuid4()),
                        "tecnico_id": tec["id"],
                        "sku": item.sku,
                        "descripcion": desc,
                        "cantidad": -item.cantidad,
                        "created_at": now_iso(),
                        "updated_at": now_iso(),
                    })
            await consumos_log.insert_one({
                "id": str(uuid.uuid4()),
                "tecnico_id": tec["id"],
                "orden_id": orden_id,
                "items": [it.model_dump() for it in materiales],
                "tipo": "pinpad_extra",
                "fecha": now_iso(),
            })
        except Exception as e:
            logger.warning("[Consumo pinpad-extra] error: %s", e)

    new_pp = {
        "id": str(uuid.uuid4()),
        "ddll": ddll_clean,
        "serie": payload.serie,
        "modelo": payload.modelo,
        "completed": True,
        "evidencia_base64": compress_evidencia(payload.evidencia_base64),
        "notas": payload.notas,
        "completed_at": now_iso(),
        "uploaded_at": now_iso(),
        "materiales_usados": [m.model_dump() for m in materiales],
        "sin_suministros": bool(payload.sin_suministros) and not materiales,
        "extra": True,  # Marca para distinguir DDLL agregadas por el técnico
        "agregado_por": tec["id"],
    }
    pin_pads.append(new_pp)
    await ordenes_col.update_one({"id": orden_id}, {"$set": {"pin_pads": pin_pads}})
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


# ----------------- Reagendar visita -----------------

class ReagendarPayload(BaseModel):
    motivo: str = Field(min_length=3)
    nueva_fecha: Optional[str] = None  # ISO datetime
    nota: Optional[str] = None


@api_router.post("/tecnico/ordenes/{orden_id}/reagendar")
async def reagendar_orden(
    orden_id: str,
    payload: ReagendarPayload,
    tec: dict = Depends(require_tecnico),
):
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    if o.get("estado") == "finalizada":
        raise HTTPException(status_code=400, detail="Orden ya finalizada")

    historial = o.get("reagendamientos") or []
    historial.append({
        "id": str(uuid.uuid4()),
        "motivo": payload.motivo,
        "nueva_fecha": payload.nueva_fecha,
        "nota": payload.nota,
        "tecnico_id": tec["id"],
        "tecnico_nombre": f"{tec.get('nombre','')} {tec.get('apellidos','')}".strip(),
        "fecha": now_iso(),
    })

    update_data = {
        "estado": "reagendada",
        "reagendamientos": historial,
        "ultimo_motivo_reagenda": payload.motivo,
    }
    if payload.nueva_fecha:
        update_data["fecha_limite"] = payload.nueva_fecha
    await ordenes_col.update_one({"id": orden_id}, {"$set": update_data})
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


# ----------------- Ruta optimizada y jornada del técnico -----------------

def _comuna_match_score(c1: Optional[str], c2: Optional[str]) -> int:
    """Score 0 (igual) → mayor (lejos). Misma comuna = 0, distinta = 1."""
    if not c1 or not c2:
        return 99
    return 0 if c1.strip().lower() == c2.strip().lower() else 1


@api_router.get("/tecnico/ruta")
async def tecnico_ruta(tec: dict = Depends(require_tecnico)):
    """Devuelve órdenes pendientes/en progreso del técnico ordenadas por
    proximidad real (nearest-neighbor) desde el domicilio del técnico, con
    fallback por región → comuna → dirección cuando faltan coordenadas.

    Estrategia de ordenamiento:
      1) Si TODAS las órdenes tienen lat/lng y el técnico también → aplicar
         nearest-neighbor puro empezando desde el domicilio del técnico.
      2) Si faltan algunas coords → priorizar primero las que coincidan con
         región y comuna del técnico, luego nearest-neighbor entre las que
         tengan coords, y al final las que no tienen coords ordenadas por
         región/comuna/dirección.
    """
    tec_comuna = (tec.get("comuna") or "").strip()
    tec_region = (tec.get("region") or "").strip()
    tec_lat = tec.get("lat")
    tec_lng = tec.get("lng")

    ordenes = await ordenes_col.find(
        {
            "tecnico_id": tec["id"],
            "estado": {"$in": ["pendiente", "en_progreso", "reagendada"]},
        },
        {"_id": 0},
    ).to_list(500)

    enriched = [await enrich_orden(o) for o in ordenes]

    def comuna_of(o):
        return ((o.get("sucursal") or {}).get("comuna") or "").strip()

    def region_of(o):
        return ((o.get("sucursal") or {}).get("region") or "").strip()

    def direccion_of(o):
        return ((o.get("sucursal") or {}).get("direccion") or "").strip()

    def latlng_of(o):
        suc = o.get("sucursal") or {}
        lat = suc.get("lat")
        lng = suc.get("lng")
        if lat is not None and lng is not None:
            try:
                return float(lat), float(lng)
            except (TypeError, ValueError):
                return None
        return None

    # Si el técnico tiene coords + hay órdenes con coords → nearest-neighbor
    if tec_lat is not None and tec_lng is not None:
        # Partir las órdenes en 4 buckets para ordenar EXACTAMENTE como pide:
        #  Bucket A: misma REGIÓN y misma COMUNA del técnico
        #  Bucket B: misma REGIÓN (otra comuna)
        #  Bucket C: otras regiones
        #  Bucket D: sin región/comuna (al final)
        same_region_same_comuna = []
        same_region_other = []
        other_region = []
        no_region = []
        for o in enriched:
            r = region_of(o).lower()
            c = comuna_of(o).lower()
            if not r:
                no_region.append(o)
            elif r == tec_region.lower() and c == tec_comuna.lower():
                same_region_same_comuna.append(o)
            elif r == tec_region.lower():
                same_region_other.append(o)
            else:
                other_region.append(o)

        # Helper: ordenar un bucket via nearest-neighbor (con coords) + fallback
        # alfabético (sin coords) intercalado al final
        def _enrich_with_coords(bucket):
            out = []
            for o in bucket:
                ll = latlng_of(o)
                if ll:
                    out.append({**o, "_lat": ll[0], "_lng": ll[1]})
                else:
                    out.append(o)
            return out

        def _sort_nn_bucket(bucket, start):
            with_coords = [
                o for o in bucket
                if o.get("_lat") is not None and o.get("_lng") is not None
            ]
            without = [
                o for o in bucket
                if o.get("_lat") is None or o.get("_lng") is None
            ]
            sorted_with = nearest_neighbor_sort(
                start, with_coords, lat_key="_lat", lng_key="_lng",
            )
            # ordenar las sin coords alfabéticamente por (region, comuna, direccion)
            without.sort(key=lambda o: (region_of(o), comuna_of(o), direccion_of(o)))
            return sorted_with + without, (sorted_with[-1] if sorted_with else None)

        cursor = (tec_lat, tec_lng)
        ordered: list = []
        for bucket in (
            same_region_same_comuna,
            same_region_other,
            other_region,
            no_region,
        ):
            bucket_e = _enrich_with_coords(bucket)
            sorted_bucket, last_in_bucket = _sort_nn_bucket(bucket_e, cursor)
            if last_in_bucket is not None:
                cursor = (last_in_bucket["_lat"], last_in_bucket["_lng"])
            ordered.extend(sorted_bucket)

        # Quitar campos auxiliares
        enriched = [
            {k: v for k, v in o.items() if k not in ("_lat", "_lng")}
            for o in ordered
        ]
    else:
        # Sin coords del técnico → fallback alfabético por región/comuna/dirección,
        # priorizando coincidencia exacta con la comuna/región del técnico
        def fallback_key(o):
            r = region_of(o).lower()
            c = comuna_of(o).lower()
            score_region = 0 if r == tec_region.lower() and tec_region else 1
            score_comuna = 0 if c == tec_comuna.lower() and tec_comuna else 1
            return (score_region, score_comuna, r, c, direccion_of(o).lower())
        enriched.sort(key=fallback_key)

    # Limit 25 (tope diario)
    MAX_ORDERS_DAY = 25
    ordenes_dia = enriched[:MAX_ORDERS_DAY]
    ordenes_resto = enriched[MAX_ORDERS_DAY:]

    # Calcular jornada estimada
    PIN_PAD_MIN = 10
    pinpads_total = 0
    for o in ordenes_dia:
        for pp in (o.get("pin_pads") or []):
            if not pp.get("completed"):
                pinpads_total += 1
    minutos_estimados = pinpads_total * PIN_PAD_MIN
    hora_inicio = "09:00"
    h_ini, m_ini = 9, 0
    total_m = m_ini + minutos_estimados
    h_fin = (h_ini + total_m // 60) % 24
    m_fin = total_m % 60
    hora_termino = f"{h_fin:02d}:{m_fin:02d}"

    return {
        "tecnico_comuna": tec_comuna,
        "tecnico_region": tec_region,
        "tecnico_direccion": tec.get("direccion") or "",
        "tecnico_lat": tec_lat,
        "tecnico_lng": tec_lng,
        "ordenes_dia": ordenes_dia,
        "ordenes_resto": ordenes_resto,
        "max_dia": MAX_ORDERS_DAY,
        "pin_pads_pendientes_dia": pinpads_total,
        "min_estimados_jornada": minutos_estimados,
        "hora_inicio_sugerida": hora_inicio,
        "hora_termino_estimada": hora_termino,
        "minutos_por_pinpad": PIN_PAD_MIN,
    }


# ----------------- Re-geocode masivo (admin) -----------------

@api_router.post("/admin/geocode/sucursales")
async def regeocode_sucursales(
    limit: int = Query(50, ge=1, le=200),
    only_missing: bool = Query(True),
    _: dict = Depends(require_admin),
):
    """Geocodifica sucursales que aún no tengan lat/lng (o todas si
    ``only_missing=false``). Devuelve resumen. Usar con calma porque Nominatim
    tiene rate limit de ~1 req/s.
    """
    query = {}
    if only_missing:
        query = {"$or": [{"lat": None}, {"lat": {"$exists": False}}]}
    cursor = sucursales_col.find(query).limit(limit)
    docs = await cursor.to_list(limit)
    ok = 0
    fail = 0
    for s in docs:
        if not s.get("direccion"):
            fail += 1
            continue
        try:
            r = await geocode_address(
                db, s.get("direccion"), s.get("comuna"), s.get("region")
            )
            if r:
                lat, lng, _disp = r
                await sucursales_col.update_one(
                    {"id": s["id"]},
                    {"$set": {"lat": lat, "lng": lng}},
                )
                ok += 1
            else:
                fail += 1
        except Exception as e:
            logger.warning("[regeocode] %s err: %s", s.get("direccion"), e)
            fail += 1
    return {"ok": ok, "fail": fail, "procesadas": len(docs)}


# ----------------- Asignación masiva (admin) -----------------

class AsignacionBulkPayload(BaseModel):
    orden_ids: List[str]
    tecnico_id: str


@api_router.post("/admin/ordenes/asignar-bulk")
async def asignacion_bulk(
    payload: AsignacionBulkPayload, _: dict = Depends(require_admin)
):
    """Asigna explícitamente N órdenes seleccionadas a UN técnico específico."""
    if not payload.orden_ids:
        raise HTTPException(status_code=400, detail="Selecciona al menos 1 orden")
    tec = await users_col.find_one(
        {"id": payload.tecnico_id, "role": "tecnico"}, {"_id": 0}
    )
    if not tec:
        raise HTTPException(status_code=404, detail="Técnico no encontrado")
    result = await ordenes_col.update_many(
        {"id": {"$in": payload.orden_ids}},
        {"$set": {"tecnico_id": payload.tecnico_id}},
    )
    # Enviar WhatsApp por cada orden (best-effort)
    enviados = 0
    for oid in payload.orden_ids:
        try:
            await _send_assignment_whatsapp(oid)
            enviados += 1
        except Exception as e:
            logger.warning("[asignacion_bulk] WhatsApp orden=%s err=%s", oid, e)
    return {
        "asignadas": result.modified_count,
        "whatsapps_enviados": enviados,
        "tecnico": f"{tec.get('nombre','')} {tec.get('apellidos','')}".strip(),
    }


class AsignacionMasivaPayload(BaseModel):
    orden_ids: Optional[List[str]] = None  # Si null → todas las pendientes sin asignar
    tecnico_ids: Optional[List[str]] = None  # Si null → todos los técnicos
    max_por_tecnico: int = 25


@api_router.post("/admin/ordenes/asignar-masivo")
async def asignacion_masiva(
    payload: AsignacionMasivaPayload, _: dict = Depends(require_admin)
):
    """Distribuye órdenes pendientes entre técnicos basándose en cercanía
    (comuna) del técnico al comercio. Máx max_por_tecnico órdenes/téc."""
    # Cargar técnicos
    if payload.tecnico_ids:
        tec_query = {"role": "tecnico", "id": {"$in": payload.tecnico_ids}}
    else:
        tec_query = {"role": "tecnico"}
    tecnicos = await users_col.find(tec_query, {"_id": 0, "hashed_password": 0}).to_list(500)
    if not tecnicos:
        raise HTTPException(status_code=400, detail="No hay técnicos disponibles")

    # Cargar órdenes
    if payload.orden_ids:
        ord_query = {"id": {"$in": payload.orden_ids}}
    else:
        ord_query = {"estado": "pendiente", "tecnico_id": {"$in": [None, ""]}}
    ordenes = await ordenes_col.find(ord_query, {"_id": 0}).to_list(2000)
    if not ordenes:
        return {"asignadas": 0, "detalle": [], "mensaje": "No hay órdenes para asignar"}

    # Resolver comuna por orden (via sucursal)
    sucursal_ids = list({o.get("sucursal_id") for o in ordenes if o.get("sucursal_id")})
    sucursales_map = {}
    if sucursal_ids:
        sucs = await sucursales_col.find(
            {"id": {"$in": sucursal_ids}}, {"_id": 0}
        ).to_list(2000)
        sucursales_map = {s["id"]: s for s in sucs}

    # Contar carga actual por técnico (órdenes en pendiente/en_progreso)
    carga_actual = {}
    for t in tecnicos:
        carga_actual[t["id"]] = await ordenes_col.count_documents(
            {"tecnico_id": t["id"], "estado": {"$in": ["pendiente", "en_progreso"]}}
        )

    asignaciones = []
    for o in ordenes:
        suc = sucursales_map.get(o.get("sucursal_id"), {})
        orden_comuna = (suc.get("comuna") or "").strip()
        # Elegir técnico con menor score de distancia y carga disponible
        candidatos = sorted(
            tecnicos,
            key=lambda t: (
                _comuna_match_score(t.get("comuna"), orden_comuna),
                carga_actual.get(t["id"], 0),
            ),
        )
        elegido = None
        for t in candidatos:
            if carga_actual.get(t["id"], 0) < payload.max_por_tecnico:
                elegido = t
                break
        if not elegido:
            continue
        await ordenes_col.update_one(
            {"id": o["id"]},
            {"$set": {"tecnico_id": elegido["id"], "estado": "pendiente"}},
        )
        carga_actual[elegido["id"]] = carga_actual.get(elegido["id"], 0) + 1
        asignaciones.append({
            "orden_id": o["id"],
            "orden_numero": o.get("numero"),
            "tecnico_id": elegido["id"],
            "tecnico_nombre": f"{elegido.get('nombre','')} {elegido.get('apellidos','')}".strip(),
            "comuna_orden": orden_comuna,
            "comuna_tecnico": elegido.get("comuna") or "",
        })
        # Enviar WhatsApp en background
        try:
            await _send_assignment_whatsapp(o["id"])
        except Exception as e:
            logger.warning("[asignacion_masiva] WhatsApp error orden=%s: %s", o["id"], e)

    return {
        "asignadas": len(asignaciones),
        "detalle": asignaciones,
        "carga_final": carga_actual,
    }


# ============================================================
#   DISPONIBILIDAD SEMANAL DEL TÉCNICO
# ============================================================

_DEFAULT_DISPONIBILIDAD = {
    d: {"activo": False, "hora_inicio": "09:00", "hora_fin": "20:00"}
    for d in ("lun", "mar", "mie", "jue", "vie", "sab", "dom")
}


@api_router.get("/tecnico/disponibilidad")
async def get_mi_disponibilidad(tec: dict = Depends(require_tecnico)):
    """Devuelve la disponibilidad semanal del técnico autenticado."""
    user = await users_col.find_one({"id": tec["id"]}, {"_id": 0, "hashed_password": 0})
    disp = (user or {}).get("disponibilidad") or _DEFAULT_DISPONIBILIDAD
    return {"disponibilidad": disp}


@api_router.put("/tecnico/disponibilidad")
async def set_mi_disponibilidad(
    payload: DisponibilidadSemanal, tec: dict = Depends(require_tecnico)
):
    """El técnico configura su disponibilidad semanal."""
    data = payload.model_dump()
    # Validación básica de formato HH:MM
    for dia, info in data.items():
        for k in ("hora_inicio", "hora_fin"):
            v = info.get(k, "")
            if not (isinstance(v, str) and len(v) == 5 and v[2] == ":"):
                raise HTTPException(
                    status_code=400,
                    detail=f"Formato hora inválido en {dia}.{k} (esperado HH:MM)",
                )
        if info["activo"] and info["hora_fin"] <= info["hora_inicio"]:
            raise HTTPException(
                status_code=400,
                detail=f"En {dia} la hora_fin debe ser mayor que hora_inicio",
            )
    await users_col.update_one(
        {"id": tec["id"]}, {"$set": {"disponibilidad": data, "disponibilidad_updated_at": now_iso()}}
    )
    return {"ok": True, "disponibilidad": data}


@api_router.get("/admin/disponibilidad")
async def admin_get_disponibilidad(_: dict = Depends(require_admin)):
    """Admin: ve la disponibilidad de TODOS los técnicos."""
    tecs = await users_col.find(
        {"role": "tecnico"},
        {"_id": 0, "hashed_password": 0},
    ).to_list(500)
    out = []
    for t in tecs:
        disp = t.get("disponibilidad") or _DEFAULT_DISPONIBILIDAD
        out.append({
            "id": t["id"],
            "nombre": t.get("nombre", ""),
            "apellidos": t.get("apellidos", ""),
            "email": t.get("email", ""),
            "telefono": t.get("telefono", ""),
            "comuna": t.get("comuna", ""),
            "region": t.get("region", ""),
            "disponibilidad": disp,
            "disponibilidad_updated_at": t.get("disponibilidad_updated_at"),
        })
    out.sort(key=lambda x: (x["apellidos"], x["nombre"]))
    return out


# ============================================================
#   CUE (Comprobante Único Electrónico) por orden
# ============================================================

@api_router.post("/tecnico/ordenes/{orden_id}/cue")
async def tecnico_upload_cue(
    orden_id: str,
    payload: OrdenCUEUpload,
    tec: dict = Depends(require_tecnico),
):
    """El técnico sube la foto del CUE (Comprobante Único Electrónico) a la orden."""
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    if not payload.cue_base64:
        raise HTTPException(status_code=400, detail="Foto CUE requerida")
    compressed = compress_evidencia(payload.cue_base64)
    set_data = {
        "cue_base64": compressed,
        "cue_uploaded_at": now_iso(),
        "cue_uploaded_by": tec["id"],
    }
    if payload.notas is not None:
        set_data["cue_notas"] = payload.notas
    await ordenes_col.update_one({"id": orden_id}, {"$set": set_data})
    o = await ordenes_col.find_one({"id": orden_id}, {"_id": 0})
    return await enrich_orden(o)


@api_router.delete("/tecnico/ordenes/{orden_id}/cue")
async def tecnico_delete_cue(
    orden_id: str, tec: dict = Depends(require_tecnico)
):
    """El técnico elimina la foto del CUE (ej. cargó la equivocada)."""
    o = await ordenes_col.find_one({"id": orden_id})
    if not o:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if o.get("tecnico_id") != tec["id"]:
        raise HTTPException(status_code=403, detail="Sin acceso a esta orden")
    await ordenes_col.update_one(
        {"id": orden_id},
        {"$unset": {"cue_base64": "", "cue_notas": "", "cue_uploaded_at": "", "cue_uploaded_by": ""}},
    )
    return {"ok": True}


# Mount suministros sub-router (with /api prefix to match the main api_router)
_sum_router = build_suministros_router(db, require_admin, get_current_user)
app.include_router(_sum_router, prefix="/api")

# Include router (must be after all @api_router decorators so every route is registered)
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ----------------- Startup: seed admin -----------------
@app.on_event("startup")
async def on_startup():
    await users_col.create_index("email", unique=True)
    await users_col.create_index("rut", unique=False, sparse=True)
    await ordenes_col.create_index("created_at")
    existing = await users_col.find_one({"email": ADMIN_EMAIL.lower()})
    if not existing:
        admin_doc = {
            "id": str(uuid.uuid4()),
            "email": ADMIN_EMAIL.lower(),
            "nombre": ADMIN_NOMBRE,
            "apellidos": ADMIN_APELLIDOS,
            "role": "admin",
            "hashed_password": hash_password(ADMIN_PASSWORD),
            "created_at": now_iso(),
        }
        await users_col.insert_one(admin_doc)
        logger.info(f"Admin seed creado: {ADMIN_EMAIL}")
    else:
        logger.info(f"Admin ya existe: {ADMIN_EMAIL}")

    # Auto-cleanup: delete ordenes older than 40 days on startup
    try:
        cutoff = datetime.now(timezone.utc) - timedelta(days=40)
        r = await ordenes_col.delete_many({"created_at": {"$lt": cutoff.isoformat()}})
        if r.deleted_count:
            logger.info(
                f"[Startup cleanup] {r.deleted_count} órdenes > 40 días eliminadas"
            )
    except Exception as e:
        logger.warning(f"Cleanup error: {e}")

    # Seed suministros catalog + bodegas + config singleton
    try:
        await init_suministros(db)
    except Exception as e:
        logger.warning(f"Suministros init error: {e}")


@app.on_event("shutdown")
async def shutdown():
    client.close()
